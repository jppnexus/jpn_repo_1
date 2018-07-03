from openpyxl import Workbook

wb = Workbook()
ws1 = wb.create_sheet()
ws1.title = "worksheet1"

filepath = 'bal.xlsx'

c = ws1['A4']

ws1['A4'] = 99999

#cell_range = ws1['A1':'C2']

#for row in ws1.iter_rows('A1:C2'):
#    for cell in row:
#        print (cell)

#wb.save('balances.xlsx')
wb.save(filepath)